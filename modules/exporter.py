import io
from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config
from modules import db

GREEN = "4CAF50"
DARK_GREEN = "388E3C"
LIGHT_GREEN = "E8F5E9"
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FONT = Font(bold=True)
THIN = Side(style="thin")
BORDER = Border(bottom=Side(style="medium"))


def _month_key(date_str):
    return date_str[:7] if date_str and len(date_str) >= 7 else None


def generate_excel():
    data = db.get_receipts_for_export()
    wb = Workbook()
    wb.remove(wb.active)

    _build_summary_sheet(wb, data)
    _build_all_receipts_sheet(wb, data)
    _build_monthly_sheets(wb, data)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header_fill():
    return PatternFill("solid", fgColor=GREEN)


def _total_fill():
    return PatternFill("solid", fgColor=LIGHT_GREEN)


def _apply_header_row(ws, row_num, values):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = HEADER_FONT
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)


def _build_summary_sheet(wb, data):
    ws = wb.create_sheet("Summary")
    _apply_header_row(ws, 1, ["Month", "Year", "Receipts", "Total Spend"])

    monthly = defaultdict(lambda: {"count": 0, "total": 0.0})
    for receipt, _ in data:
        key = _month_key(receipt.get("receipt_date"))
        if key:
            monthly[key]["count"] += 1
            monthly[key]["total"] += receipt.get("total_amount") or 0.0

    keys = sorted(monthly.keys(), reverse=True)
    grand_count = 0
    grand_total = 0.0

    for row_num, key in enumerate(keys, 2):
        year, month = key.split("-")
        month_name = date(int(year), int(month), 1).strftime("%B")
        ws.cell(row=row_num, column=1, value=month_name)
        ws.cell(row=row_num, column=2, value=int(year))
        ws.cell(row=row_num, column=3, value=monthly[key]["count"])
        amount_cell = ws.cell(row=row_num, column=4, value=monthly[key]["total"])
        amount_cell.number_format = '"$"#,##0.00'
        grand_count += monthly[key]["count"]
        grand_total += monthly[key]["total"]

    total_row = len(keys) + 2
    for col, val in enumerate(["TOTAL", "", grand_count, grand_total], 1):
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font = TOTAL_FONT
        cell.fill = _total_fill()
        if col == 4:
            cell.number_format = '"$"#,##0.00'

    receipts_without_date = sum(1 for r, _ in data if not r.get("receipt_date"))
    if receipts_without_date:
        ws.cell(row=total_row + 2, column=1,
                value=f"Note: {receipts_without_date} receipt(s) have no date and are excluded from monthly totals.")

    _autofit(ws)


def _build_all_receipts_sheet(wb, data):
    ws = wb.create_sheet("All Receipts")
    _apply_header_row(ws, 1, ["Date", "Company", "Total Amount", "Filename"])

    for row_num, (receipt, _) in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=receipt.get("receipt_date") or "Unknown")
        ws.cell(row=row_num, column=2, value=receipt.get("company_name") or "Unknown")
        amount_cell = ws.cell(row=row_num, column=3, value=receipt.get("total_amount"))
        amount_cell.number_format = '"$"#,##0.00'
        ws.cell(row=row_num, column=4, value=receipt.get("filename"))

    total_row = len(data) + 2
    total = sum((r.get("total_amount") or 0) for r, _ in data)
    ws.cell(row=total_row, column=2, value="TOTAL").font = TOTAL_FONT
    amount_cell = ws.cell(row=total_row, column=3, value=total)
    amount_cell.font = TOTAL_FONT
    amount_cell.fill = _total_fill()
    amount_cell.number_format = '"$"#,##0.00'

    _autofit(ws)


def _build_monthly_sheets(wb, data):
    by_month = defaultdict(list)
    historical = []

    for receipt, items in data:
        key = _month_key(receipt.get("receipt_date"))
        if key:
            by_month[key].append((receipt, items))
        else:
            historical.append((receipt, items))

    sorted_months = sorted(by_month.keys(), reverse=True)
    recent_months = sorted_months[:24]
    old_months = sorted_months[24:]

    for month_key in recent_months:
        _build_month_sheet(wb, month_key, by_month[month_key])

    if old_months or historical:
        old_data = []
        for m in old_months:
            old_data.extend(by_month[m])
        old_data.extend(historical)
        _build_month_sheet(wb, "Historical", old_data)


def _build_month_sheet(wb, title, records):
    ws = wb.create_sheet(title)
    _apply_header_row(ws, 1, ["Date", "Company", "Item Name", "Qty", "Unit Price", "Line Total"])

    row_num = 2
    for receipt, items in records:
        date_val = receipt.get("receipt_date") or "Unknown"
        company = receipt.get("company_name") or "Unknown"

        if items:
            for item in items:
                ws.cell(row=row_num, column=1, value=date_val)
                ws.cell(row=row_num, column=2, value=company)
                ws.cell(row=row_num, column=3, value=item.get("item_name"))
                ws.cell(row=row_num, column=4, value=item.get("quantity"))
                if item.get("unit_price") is not None:
                    c = ws.cell(row=row_num, column=5, value=item.get("unit_price"))
                    c.number_format = '"$"#,##0.00'
                if item.get("line_total") is not None:
                    c = ws.cell(row=row_num, column=6, value=item.get("line_total"))
                    c.number_format = '"$"#,##0.00'
                row_num += 1
        else:
            # No line items — write a single summary row
            ws.cell(row=row_num, column=1, value=date_val)
            ws.cell(row=row_num, column=2, value=company)
            ws.cell(row=row_num, column=3, value="(no line items)")
            if receipt.get("total_amount") is not None:
                c = ws.cell(row=row_num, column=6, value=receipt.get("total_amount"))
                c.number_format = '"$"#,##0.00'
            row_num += 1

    _autofit(ws)
